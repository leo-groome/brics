"""Extracción bruta de XLSX/XLS a texto plano.

Filosofía: cero heurísticas por archivo. Solo bytes → texto.
`data_only=True` mata fórmulas y deja valores calculados.
El normalizer LLM (siguiente paso) recibe este dump y devuelve conceptos canónicos.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
import xlrd

logger = logging.getLogger(__name__)


@dataclass
class SheetDump:
    filename: str
    sheet_name: str
    raw_text: str
    n_rows: int
    n_cols: int


@dataclass
class FileDump:
    filename: str
    sheets: list[SheetDump] = field(default_factory=list)
    error: str | None = None


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        # Sin notación científica, sin .0 colgante para enteros.
        if value.is_integer():
            return str(int(value))
        return f"{value:.6g}"
    return str(value).strip()


def _rows_to_text(rows: list[list[str]]) -> str:
    """Pipe-delimited. Filas totalmente vacías se omiten."""
    lines = []
    for row in rows:
        if not any(cell for cell in row):
            continue
        lines.append(" | ".join(row))
    return "\n".join(lines)


def _extract_xlsx(path: Path) -> list[SheetDump]:
    wb = load_workbook(path, data_only=True, read_only=True)
    dumps: list[SheetDump] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            max_cols = 0
            for row in ws.iter_rows(values_only=True):
                cells = [_cell_to_str(v) for v in row]
                # Trim trailing empties para no inflar el texto.
                while cells and cells[-1] == "":
                    cells.pop()
                if cells:
                    rows.append(cells)
                    if len(cells) > max_cols:
                        max_cols = len(cells)
            raw_text = _rows_to_text(rows)
            dumps.append(
                SheetDump(
                    filename=path.name,
                    sheet_name=sheet_name,
                    raw_text=raw_text,
                    n_rows=len(rows),
                    n_cols=max_cols,
                )
            )
    finally:
        wb.close()
    return dumps


def _extract_xls(path: Path) -> list[SheetDump]:
    book = xlrd.open_workbook(str(path), formatting_info=False, on_demand=True)
    dumps: list[SheetDump] = []
    try:
        for sheet_idx in range(book.nsheets):
            sh = book.sheet_by_index(sheet_idx)
            rows: list[list[str]] = []
            max_cols = 0
            for r in range(sh.nrows):
                cells = [_cell_to_str(sh.cell_value(r, c)) for c in range(sh.ncols)]
                while cells and cells[-1] == "":
                    cells.pop()
                if cells:
                    rows.append(cells)
                    if len(cells) > max_cols:
                        max_cols = len(cells)
            raw_text = _rows_to_text(rows)
            dumps.append(
                SheetDump(
                    filename=path.name,
                    sheet_name=sh.name,
                    raw_text=raw_text,
                    n_rows=len(rows),
                    n_cols=max_cols,
                )
            )
            book.unload_sheet(sheet_idx)
    finally:
        book.release_resources()
    return dumps


def extract_file(path: Path) -> FileDump:
    suffix = path.suffix.lower()
    try:
        if suffix == ".xlsx":
            sheets = _extract_xlsx(path)
        elif suffix == ".xls":
            sheets = _extract_xls(path)
        else:
            return FileDump(filename=path.name, error=f"unsupported extension: {suffix}")
        return FileDump(filename=path.name, sheets=sheets)
    except Exception as e:  # noqa: BLE001
        logger.exception("extract_file failed: %s", path.name)
        return FileDump(filename=path.name, error=str(e))


def extract_dir(data_dir: Path) -> list[FileDump]:
    files = sorted(p for p in data_dir.iterdir() if p.suffix.lower() in {".xlsx", ".xls"})
    return [extract_file(p) for p in files]
